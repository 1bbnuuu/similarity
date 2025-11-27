import requests
from bs4 import BeautifulSoup
import time
from urllib.parse import urljoin
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

BASE_URL = "https://perpus.stmikplk.ac.id"

URLS_TO_SCRAPE = [
    f"{BASE_URL}/perpus/main/tipeItem/21?max=10&offset=0",
    f"{BASE_URL}/perpus/main/tipeItem/24?max=10&offset=0",
    f"{BASE_URL}/perpus/main/tipeItem/4?max=10&offset=0",
    f"{BASE_URL}/perpus/main/tipeItem/17?max=10&offset=0",
]

DUPLICATE_THRESHOLD = 2

def scrape_page(url):
    print(f"\nScraping: {url}")
    
    try:
        response = requests.get(url, timeout=10)
        response.encoding = 'utf-8'
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        items = []
        columns = soup.find_all('div', class_='column nine last')
        
        for col in columns:
            h2 = col.find('h2')
            if not h2:
                continue
            
            link_tag = h2.find('a', href=True)
            if not link_tag:
                continue
            
            item_link = urljoin(BASE_URL, link_tag['href'])
            
            nim = ''
            h5 = col.find('h5', class_='meta-post')
            if h5:
                next_p = h5.find_next_sibling('p')
                if next_p:
                    for content in next_p.contents:
                        if isinstance(content, str):
                            nim = content.strip()
                            break

            ul = col.find('ul', style=lambda x: x and 'font-weight: bold' in x)
            
            if not ul:
                continue
            
            judul = ''
            pengarang = ''
            tahun = ''
            
            for li in ul.find_all('li'):
                text = li.get_text(strip=True)
                
                if text.startswith('Judul :'):
                    link = li.find('a')
                    judul = link.get_text(strip=True) if link else ''
                
                elif text.startswith('Pengarang :'):
                    authors = [a.get_text(strip=True) for a in li.find_all('a')]
                    pengarang = ', '.join(authors) if authors else ''
                
                elif text.startswith('Tahun :'):
                    link = li.find('a')
                    tahun = link.get_text(strip=True) if link else ''
            
            if judul:
                items.append({
                    'judul': judul,
                    'nim': nim,
                    'tahun': tahun,
                    'penulis': pengarang,
                    'file': item_link
                })
                print(f"  {judul[:50]}...")
        
        pagination = soup.find('div', class_='pagination')
        next_url = None
        if pagination:
            next_link = pagination.find('a', class_='nextLink')
            if next_link and next_link.get('href'):
                next_url = urljoin(BASE_URL, next_link['href'])
        
        return items, next_url
    
    except Exception as e:
        print(f"  Error: {e}")
        return [], None

def load_existing_urls(filename):
    if not os.path.exists(filename):
        return set()
    
    existing_urls = set()
    try:
        wb = load_workbook(filename)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        if 'file' in headers:
            file_col = headers.index('file') + 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[file_col - 1]:
                    existing_urls.add(row[file_col - 1])
        
        wb.close()
    except Exception as e:
        print(f"Error membaca file: {e}")
        return set()
    
    return existing_urls

def scrape_all(start_url, seen_urls, max_pages=None):
    all_data = []
    current_url = start_url
    page = 0
    consecutive_duplicates = 0
    
    while current_url:
        page += 1
        print(f"\nHalaman {page}")
        
        items, next_url = scrape_page(current_url)
        
        if not items:
            print("  Tidak ada item ditemukan")
            break
        
        new_items_in_page = 0
        duplicate_items_in_page = 0
        
        for item in items:
            if item['file'] not in seen_urls:
                all_data.append(item)
                seen_urls.add(item['file'])
                new_items_in_page += 1
                consecutive_duplicates = 0
            else:
                print(f"  [SKIP] Duplikat: {item['judul'][:40]}...")
                duplicate_items_in_page += 1
                consecutive_duplicates += 1
        
        print(f"  Baru: {new_items_in_page}, Duplikat: {duplicate_items_in_page}")
        
        if new_items_in_page == 0:
            print(f"\nSemua data di halaman ini sudah ada")
            break
        
        if consecutive_duplicates >= DUPLICATE_THRESHOLD:
            break
        
        if max_pages and page >= max_pages:
            print(f"\nBatas {max_pages} halaman tercapai")
            break
        
        if not next_url:
            print("\nTidak ada halaman berikutnya")
            break
        
        current_url = next_url
        time.sleep(1)
    
    return all_data

def save_to_xlsx(data, filename):
    if not data:
        print("Tidak ada data baru untuk disimpan")
        return
    
    directory = os.path.dirname(filename)
    if directory and not os.path.exists(directory):
        os.makedirs(directory, exist_ok=True)
    
    file_exists = os.path.exists(filename)
    
    try:
        if file_exists:
            wb = load_workbook(filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Scraping Results"
            
            # Tulis header
            headers = ['Judul', 'NIM', 'Tahun', 'Penulis', 'File']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        
        # Tambahkan data baru
        start_row = ws.max_row + 1
        for idx, item in enumerate(data, start_row):
            ws.cell(row=idx, column=1, value=item['judul'])
            ws.cell(row=idx, column=2, value=item['nim'])
            ws.cell(row=idx, column=3, value=item['tahun'])
            ws.cell(row=idx, column=4, value=item['penulis'])
            ws.cell(row=idx, column=5, value=item['file'])
        
        wb.save(filename)
        wb.close()
        
        print(f"\n{len(data)} data baru disimpan ke {filename}")
    except Exception as e:
        print(f"\nError menyimpan ke XLSX: {e}")

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    filename = os.path.join(script_dir, 'scraping.xlsx')
    
    seen_urls = load_existing_urls(filename)
    existing_count = len(seen_urls)
    
    print("="*60)
    print("PERPUS STMIK PLK SCRAPER")
    print("="*60)
    print(f"File output: {filename}")
    print(f"Data existing: {existing_count} record")
    print(f"Early stop threshold: {DUPLICATE_THRESHOLD} duplikat berturut-turut")
    
    total_data = 0
    
    for idx, url in enumerate(URLS_TO_SCRAPE, 1):
        print(f"\n{'='*60}")
        print(f"URL {idx}/{len(URLS_TO_SCRAPE)}")
        print(f"{'='*60}")
        
        data = scrape_all(url, seen_urls)
        
        if data:
            save_to_xlsx(data, filename)
            total_data += len(data)
            print(f"Berhasil: {len(data)} data baru dari URL {idx}")
        else:
            print(f"Tidak ada data baru dari URL {idx}")
        
        if idx < len(URLS_TO_SCRAPE):
            time.sleep(1)
    
    print(f"\n{'='*60}")
    print("SELESAI")
    print(f"Total data baru: {total_data} record")
    print(f"Total data di Excel: {existing_count + total_data} record")
    print("="*60)

if __name__ == "__main__":
    main()